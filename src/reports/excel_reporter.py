"""
Excel Report Generator
Excel 報表生成器
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelReporter:
    """Excel 報表生成器"""

    def __init__(self, output_path: Path):
        """
        初始化報表生成器

        Args:
            output_path: 輸出路徑
        """
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def generate_compliance_report(self, data: List[Dict], statistics: Dict,
                                   conflicts: List[Dict], differences: List[Dict]):
        """
        生成完整的合規報表

        Args:
            data: 統一成分資料
            statistics: 統計資訊
            conflicts: 衝突列表
            differences: 差異列表
        """
        # 創建 Excel Writer
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            # 1. 主表格 - 全部成分
            df_main = pd.DataFrame(data)
            df_main.to_excel(writer, sheet_name='Global Compliance', index=False)

            # 2. 統計摘要
            df_stats = self._create_statistics_sheet(statistics)
            df_stats.to_excel(writer, sheet_name='Statistics', index=False)

            # 3. 衝突成分
            if conflicts:
                df_conflicts = pd.DataFrame(conflicts)
                df_conflicts.to_excel(writer, sheet_name='Conflicts', index=False)

            # 4. 限量差異
            if differences:
                df_diff = pd.DataFrame(differences)
                df_diff.to_excel(writer, sheet_name='Limit Differences', index=False)

            # 5. 各地區詳細資料
            for region in ['EU', 'ASEAN', 'Japan', 'Canada', 'China']:
                df_region = self._filter_by_region(df_main, region)
                if not df_region.empty:
                    df_region.to_excel(writer, sheet_name=region, index=False)

        # 應用格式化
        self._apply_formatting()

    def _create_statistics_sheet(self, statistics: Dict) -> pd.DataFrame:
        """
        創建統計工作表

        Args:
            statistics: 統計資訊

        Returns:
            統計 DataFrame
        """
        rows = []

        # 總體統計
        rows.append(['Overall Statistics', ''])
        rows.append(['Total Ingredients', statistics.get('total_ingredients', 0)])
        rows.append(['With CAS Number', statistics.get('with_cas', 0)])
        rows.append(['Conflicts Detected', statistics.get('conflicts', 0)])
        rows.append(['Limit Differences', statistics.get('limit_differences', 0)])
        rows.append(['', ''])

        # 各地區統計
        rows.append(['By Region', 'Count'])
        by_region = statistics.get('by_region', {})
        for region, count in by_region.items():
            rows.append([region, count])

        df = pd.DataFrame(rows, columns=['Metric', 'Value'])
        return df

    def _filter_by_region(self, df: pd.DataFrame, region: str) -> pd.DataFrame:
        """
        篩選特定地區的成分

        Args:
            df: 主 DataFrame
            region: 地區名稱

        Returns:
            篩選後的 DataFrame
        """
        status_col = f'{region}_Status'

        if status_col not in df.columns:
            return pd.DataFrame()

        # 篩選出該地區有資料的成分
        df_filtered = df[df[status_col] != 'Not_Listed'].copy()

        # 只保留相關欄位
        relevant_cols = ['INCI_Name', 'CAS_No', 'Chinese_Name']
        for col in df.columns:
            if col.startswith(region):
                relevant_cols.append(col)

        df_filtered = df_filtered[relevant_cols]

        return df_filtered

    def _apply_formatting(self):
        """應用 Excel 格式化"""
        wb = load_workbook(self.output_path)

        # 定義樣式
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        prohibited_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        restricted_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        allowed_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        conflict_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 格式化每個工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 格式化標題行
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            # 自動調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 應用條件格式化
            if sheet_name == 'Global Compliance':
                self._apply_conditional_formatting(ws, prohibited_fill, restricted_fill,
                                                   allowed_fill, conflict_fill)

            # 凍結首行
            ws.freeze_panes = 'A2'

        wb.save(self.output_path)

    def _apply_conditional_formatting(self, ws, prohibited_fill, restricted_fill,
                                     allowed_fill, conflict_fill):
        """
        應用條件格式化

        Args:
            ws: 工作表物件
            prohibited_fill: 禁用填充
            restricted_fill: 限用填充
            allowed_fill: 允用填充
            conflict_fill: 衝突填充
        """
        # 找到狀態欄位
        status_columns = []
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and 'Status' in str(cell.value):
                status_columns.append(col_idx)

        # 應用格式
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in status_columns:
                cell = ws.cell(row=row_idx, column=col_idx)

                if cell.value == 'Prohibited':
                    cell.fill = prohibited_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                elif cell.value == 'Restricted':
                    cell.fill = restricted_fill
                    cell.font = Font(bold=True)
                elif cell.value == 'Allowed':
                    cell.fill = allowed_fill

            # 標記衝突行
            conflicts_cell = ws.cell(row=row_idx, column=ws.max_column)
            if conflicts_cell.value and str(conflicts_cell.value).strip():
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = conflict_fill

    def generate_simple_export(self, data: List[Dict], sheet_name: str = 'Data'):
        """
        生成簡單的 Excel 匯出

        Args:
            data: 資料列表
            sheet_name: 工作表名稱
        """
        df = pd.DataFrame(data)
        df.to_excel(self.output_path, sheet_name=sheet_name, index=False)

        # 簡單格式化
        wb = load_workbook(self.output_path)
        ws = wb[sheet_name]

        # 標題行格式
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        wb.save(self.output_path)
