"""
数据导出工具

支持将法规数据导出为多种格式：
1. Excel (.xlsx) - 多工作表格式
2. CSV - 单表格式
3. JSON - 结构化格式

用于方便数据分析和报告生成。
"""

import json
import csv
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

logger = logging.getLogger(__name__)

# 尝试导入 openpyxl（Excel 支持）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export will not be available.")


class DataExporter:
    """数据导出器"""

    def __init__(self, data_dir: Path):
        """
        初始化导出器

        Args:
            data_dir: 数据目录（包含六张表的 JSON 文件）
        """
        self.data_dir = Path(data_dir)

        if not self.data_dir.exists():
            raise ValueError(f"Data directory not found: {data_dir}")

    def load_table_data(self, table_type: str) -> Dict[str, Any]:
        """
        加载表数据

        Args:
            table_type: 表类型

        Returns:
            表数据字典
        """
        table_file = self.data_dir / f"{table_type}_latest.json"

        if not table_file.exists():
            logger.warning(f"Table file not found: {table_file}")
            return {"records": []}

        with open(table_file, 'r', encoding='utf-8') as f:
            return json.load(f)

    def export_to_excel(
        self,
        output_file: Path,
        tables: Optional[List[str]] = None
    ) -> bool:
        """
        导出为 Excel 文件（多工作表）

        Args:
            output_file: 输出文件路径
            tables: 要导出的表列表（默认全部）

        Returns:
            是否成功
        """
        if not EXCEL_AVAILABLE:
            logger.error("Excel export requires openpyxl. Install with: pip install openpyxl")
            return False

        if tables is None:
            tables = ['prohibited', 'restricted', 'preservatives', 'uv_filters', 'colorants', 'whitelist']

        logger.info(f"Exporting to Excel: {output_file}")

        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认工作表

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 为每个表创建工作表
        for table_type in tables:
            data = self.load_table_data(table_type)
            records = data.get('records', [])

            if not records:
                logger.info(f"  Skipping empty table: {table_type}")
                continue

            # 创建工作表
            ws = wb.create_sheet(title=table_type.replace('_', ' ').title())

            # 获取字段列表（从第一条记录）
            if records:
                headers = list(records[0].keys())

                # 写入表头
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 写入数据
                for row_idx, record in enumerate(records, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        value = record.get(header)
                        # 处理 None 值
                        if value is None:
                            value = ""
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # 自动调整列宽
                for col_idx, header in enumerate(headers, start=1):
                    column_letter = get_column_letter(col_idx)
                    # 设置最小宽度为 12，最大宽度为 50
                    max_length = len(str(header))
                    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = min(max(max_length + 2, 12), 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # 冻结首行
                ws.freeze_panes = 'A2'

            logger.info(f"  Added sheet: {table_type} ({len(records)} records)")

        # 添加汇总工作表
        self._add_summary_sheet(wb)

        # 保存文件
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)

        logger.info(f"Excel file saved: {output_file}")
        return True

    def _add_summary_sheet(self, wb: 'Workbook'):
        """
        添加汇总工作表

        Args:
            wb: 工作簿对象
        """
        ws = wb.create_sheet(title="Summary", index=0)

        # 标题
        ws['A1'] = "法规数据汇总"
        ws['A1'].font = Font(bold=True, size=14)

        # 生成时间
        ws['A2'] = "生成时间:"
        ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 表统计
        ws['A4'] = "表名"
        ws['B4'] = "记录数"
        ws['A4'].font = Font(bold=True)
        ws['B4'].font = Font(bold=True)

        row = 5
        for sheet in wb.sheetnames:
            if sheet == "Summary":
                continue

            sheet_obj = wb[sheet]
            record_count = sheet_obj.max_row - 1  # 减去表头

            ws[f'A{row}'] = sheet
            ws[f'B{row}'] = record_count
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def export_to_csv(
        self,
        output_dir: Path,
        tables: Optional[List[str]] = None
    ) -> bool:
        """
        导出为 CSV 文件（每个表一个文件）

        Args:
            output_dir: 输出目录
            tables: 要导出的表列表（默认全部）

        Returns:
            是否成功
        """
        if tables is None:
            tables = ['prohibited', 'restricted', 'preservatives', 'uv_filters', 'colorants', 'whitelist']

        logger.info(f"Exporting to CSV: {output_dir}")

        output_dir.mkdir(parents=True, exist_ok=True)

        # 为每个表导出 CSV
        for table_type in tables:
            data = self.load_table_data(table_type)
            records = data.get('records', [])

            if not records:
                logger.info(f"  Skipping empty table: {table_type}")
                continue

            # CSV 文件路径
            csv_file = output_dir / f"{table_type}.csv"

            # 获取字段列表
            headers = list(records[0].keys())

            # 写入 CSV
            with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(records)

            logger.info(f"  Saved: {csv_file} ({len(records)} records)")

        logger.info(f"CSV files saved to: {output_dir}")
        return True

    def export_master_view_to_excel(
        self,
        output_file: Path,
        master_view_file: Optional[Path] = None
    ) -> bool:
        """
        导出 MasterView 为 Excel（跨国比对表）

        Args:
            output_file: 输出文件路径
            master_view_file: MasterView JSON 文件路径

        Returns:
            是否成功
        """
        if not EXCEL_AVAILABLE:
            logger.error("Excel export requires openpyxl")
            return False

        # 加载 MasterView
        if master_view_file is None:
            master_view_file = self.data_dir / 'master_view.json'

        if not master_view_file.exists():
            logger.error(f"MasterView file not found: {master_view_file}")
            return False

        with open(master_view_file, 'r', encoding='utf-8') as f:
            master_view = json.load(f)

        data = master_view.get('data', [])

        if not data:
            logger.warning("MasterView is empty")
            return False

        logger.info(f"Exporting MasterView to Excel: {output_file}")

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Cross-Jurisdiction Comparison"

        # 表头
        jurisdictions = master_view.get('jurisdictions', ['EU', 'ASEAN', 'JP', 'CA', 'CN'])

        headers = ['INCI Name', 'CAS No']
        for jurisdiction in jurisdictions:
            headers.extend([
                f'{jurisdiction} Status',
                f'{jurisdiction} Max Conc (%)',
                f'{jurisdiction} Conditions'
            ])

        # 样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for row_idx, ingredient in enumerate(data, start=2):
            col_idx = 1

            # 基本信息
            ws.cell(row=row_idx, column=col_idx, value=ingredient['INCI_Name'])
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx, value=ingredient.get('CAS_No', ''))
            col_idx += 1

            # 各法规属地信息
            for jurisdiction in jurisdictions:
                regulation = ingredient['Regulations'].get(jurisdiction, {})

                status = regulation.get('Status', '未规定')
                max_conc = regulation.get('Max_Conc_Percent')
                conditions = regulation.get('Conditions', '')

                ws.cell(row=row_idx, column=col_idx, value=status)
                col_idx += 1
                ws.cell(row=row_idx, column=col_idx, value=max_conc if max_conc is not None else '')
                col_idx += 1
                ws.cell(row=row_idx, column=col_idx, value=conditions)
                col_idx += 1

        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20

        # 冻结首行和前两列
        ws.freeze_panes = 'C2'

        # 保存文件
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)

        logger.info(f"MasterView Excel saved: {output_file} ({len(data)} ingredients)")
        return True


if __name__ == "__main__":
    """测试代码"""
    import logging

    logging.basicConfig(level=logging.INFO)

    # 测试导出
    data_dir = Path("data/integrated")

    if data_dir.exists():
        exporter = DataExporter(data_dir)

        # 导出为 Excel
        exporter.export_to_excel(Path("output/regulations.xlsx"))

        # 导出为 CSV
        exporter.export_to_csv(Path("output/csv"))

        # 导出 MasterView
        exporter.export_master_view_to_excel(Path("output/cross_jurisdiction_comparison.xlsx"))

        print("\n导出完成!")
    else:
        print(f"Data directory not found: {data_dir}")
        print("Please run data integration first.")
